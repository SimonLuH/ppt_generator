# slide_mapping_editor.py

import json,os
from typing import Dict,Any,List
from PyQt5.QtWidgets import (
    QDialog,QVBoxLayout,QHBoxLayout,
    QTableWidget,QTableWidgetItem,QPushButton,
    QLabel,QComboBox,QMessageBox,QFileDialog
)
from PyQt5.QtCore import Qt

class SlideMappingEditor(QDialog):
    """
    可配置 slide_mappings,含枚举: type=[row_for_page,row_for_table_row], copy=[True,False],
    sheet从Excel获取
    """
    def __init__(self,parent=None):
        super().__init__(parent)
        self.setWindowTitle("编辑 slide_mappings")
        self.resize(600,300)
        self.table=QTableWidget(0,4,self)
        self.table.setHorizontalHeaderLabels(["Index","type","sheet","copy"])
        self.btnLoadSheet=QPushButton("从Excel取sheet")
        self.btnAdd=QPushButton("添加行")
        self.btnDel=QPushButton("删除行")
        self.btnLoadCfg=QPushButton("加载JSON")
        self.btnSaveCfg=QPushButton("保存JSON")
        lay=QVBoxLayout()
        lay.addWidget(self.table)
        h1=QHBoxLayout()
        h1.addWidget(self.btnLoadSheet)
        h1.addWidget(self.btnAdd)
        h1.addWidget(self.btnDel)
        h1.addWidget(self.btnLoadCfg)
        h1.addWidget(self.btnSaveCfg)
        lay.addLayout(h1)
        self.setLayout(lay)
        self.btnLoadSheet.clicked.connect(self.loadSheetsFromExcel)
        self.btnAdd.clicked.connect(self.addRow)
        self.btnDel.clicked.connect(self.delRow)
        self.btnLoadCfg.clicked.connect(self.loadJson)
        self.btnSaveCfg.clicked.connect(self.saveJson)
        self.sheetList=[]

    def loadSheetsFromExcel(self):
        """
        让用户选某Excel,获取sheet名 => self.sheetList
        """
        path,_=QFileDialog.getOpenFileName(self,"选Excel",".","Excel(*.xlsx)")
        if not path:return
        try:
            from openpyxl import load_workbook
            wb=load_workbook(path,read_only=True)
            self.sheetList=wb.sheetnames
            QMessageBox.information(self,"提示",f"获取sheet成功:{self.sheetList}")
        except Exception as e:
            QMessageBox.critical(self,"错误",f"读取失败:{str(e)}")

    def addRow(self):
        """
        添加一行,内含ComboBox(枚举)等
        """
        r=self.table.rowCount()
        self.table.insertRow(r)
        # index列(用户手输)
        idxItem=QTableWidgetItem("")
        self.table.setItem(r,0,idxItem)
        # type列(下拉: row_for_page / row_for_table_row)
        cbType=QComboBox();cbType.addItems(["row_for_page","row_for_table_row"])
        self.table.setCellWidget(r,1,cbType)
        # sheet列(下拉: self.sheetList)
        cbSheet=QComboBox();cbSheet.addItems(self.sheetList if self.sheetList else ["Sheet1"])
        self.table.setCellWidget(r,2,cbSheet)
        # copy列(下拉: True/False)
        cbCopy=QComboBox();cbCopy.addItems(["False","True"])
        self.table.setCellWidget(r,3,cbCopy)

    def delRow(self):
        """
        删除选中行
        """
        row=self.table.currentRow()
        if row>=0:self.table.removeRow(row)

    def loadJson(self):
        """
        从 json 文件载入,填表
        """
        p, _ = QFileDialog.getOpenFileName(self, "载入slide_mappings", ".", "JSON Files(*.json)")
        if not p: return
        try:
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
            self.table.setRowCount(0)
            for k, v in data.items():
                r = self.table.rowCount()
                self.table.insertRow(r)
                # 第0列: index
                idxItem = QTableWidgetItem(str(k))
                self.table.setItem(r, 0, idxItem)
                # 第1列: type
                cbType = QComboBox()
                cbType.addItems(["row_for_page", "row_for_table_row"])
                if "type" in v: cbType.setCurrentText(str(v["type"]))
                self.table.setCellWidget(r, 1, cbType)
                # 第2列: sheet
                cbSheet = QComboBox()
                # 先加载 self.sheetList 如果有 (若你是先 "获取sheet"那个按钮?), 否则可以临时加 v["sheet"]
                if self.sheetList:  # 如果先 loadSheetsFromExcel
                    cbSheet.addItems(self.sheetList)
                # 如果v["sheet"]不在self.sheetList,可先addItem再setCurrentText
                if "sheet" in v:
                    if self.sheetList and v["sheet"] not in self.sheetList:
                        cbSheet.addItem(v["sheet"])
                    cbSheet.setCurrentText(str(v["sheet"]))
                self.table.setCellWidget(r, 2, cbSheet)
                # 第3列: copy
                cbCopy = QComboBox()
                cbCopy.addItems(["False", "True"])
                if v.get("copy", False): cbCopy.setCurrentText("True")
                self.table.setCellWidget(r, 3, cbCopy)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取失败: {str(e)}")

    def saveJson(self):
        """
        将表里内容写json
        """
        rCount=self.table.rowCount()
        mappings={}
        for r in range(rCount):
            idxItem=self.table.item(r,0)
            if not idxItem:continue
            idx=idxItem.text().strip()
            if not idx.isdigit():continue
            k=int(idx)
            v={}
            cbT=self.table.cellWidget(r,1)
            v["type"]=cbT.currentText()
            cbS=self.table.cellWidget(r,2)
            v["sheet"]=cbS.currentText()
            cbC=self.table.cellWidget(r,3)
            if cbC.currentText()=="True":v["copy"]=True
            mappings[k]=v
        p,_=QFileDialog.getSaveFileName(self,"保存","总监.json","*.json")
        if not p:return
        try:
            with open(p,"w",encoding="utf-8") as f:
                json.dump(mappings,f,indent=2,ensure_ascii=False)
            QMessageBox.information(self,"提示",f"已保存:{p}")
        except Exception as e:
            QMessageBox.critical(self,"错误",f"保存失败:{str(e)}")