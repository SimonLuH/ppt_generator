# data_access/excel_reader.py

import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP

def round_half_up(value, ndigits=2):
    """
    传统四舍五入到 ndigits 位小数
    """
    dec = Decimal(str(value))
    quant = Decimal('1.' + '0' * ndigits)
    return float(dec.quantize(quant, rounding=ROUND_HALF_UP))

class ExcelDataProvider:
    """
    从Excel读取数据的类。
    返回 {sheetName: [ { '[A]':valA, '[B]':valB, ... }, ... ]}
    """
    def __init__(self, excel_file: str):
        if not os.path.isfile(excel_file):
            raise FileNotFoundError(f"Excel文件不存在: {excel_file}")
        self.excel_file = excel_file

    def read_data(self):
        wb = load_workbook(self.excel_file, data_only=True)
        result = {}

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            sheet_data = []

            # 从第2行开始读，第一行可能是表头
            for row_idx in range(2, max_row + 1):
                row_dict = {}
                non_empty_flag = False
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell_value = cell.value

                    col_letter = get_column_letter(col_idx)
                    ph_key = f"[{col_letter}]"

                    if cell_value is not None:
                        non_empty_flag = True
                        if isinstance(cell_value, float):
                            cell_value = round_half_up(cell_value, 2)
                        elif isinstance(cell_value, datetime):
                            cell_value = cell_value.strftime('%Y-%m-%d')
                    else:
                        cell_value = ""

                    row_dict[ph_key] = cell_value

                if non_empty_flag:
                    sheet_data.append(row_dict)

            result[sheet_name] = sheet_data

        return result